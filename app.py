import openpyxl
from datetime import datetime
from openpyxl.drawing.image import Image
import os
import win32com.client as win32

# 請求データファイルを読み込み
wb = openpyxl.load_workbook("files/invoice_data.xlsx", data_only=True)

ws = wb.active

# valuesにワークシートの全てのセルの値をリストとして取得
values = list(ws.values)

# valuesリストの長さを取得します。lastrowという変数が最後の行番号となります。
lastrow = len(values)

# 請求書テンプレートファイルを読み込み
wb = openpyxl.load_workbook("files/invoice.xlsx")

ws = wb.active

# 現在の日付を取得
current_date = datetime.now()
# 日付を請求日用に整形 
invoice_date = current_date.strftime("%Y年%m月%d日")
# 番号用に整形
year_month = current_date.strftime("%Y%m")
# 件名用に整形
invoice_month = current_date.month
# 保存先のフォルダ
output_folder = f"請求書_{current_date.strftime('%Y年%m月')}"
# フォルダ作成 exist_ok=Trueで既にディレクトリに存在してもエラーが発生しない
os.makedirs(output_folder, exist_ok=True)
# 請求書ファイルのパスを作成
output_file = f"{output_folder}/請求書_{current_date.strftime('%Y年%m月')}.xlsx"

# 請求書番号の初期化（毎月リセット）
invoice_number = 1

# 繰り返し処理でデータを取得
for index in range(lastrow):
    if not index == 0:
        if values[index][12] is None:
            continue

        sheet_name = str(values[index][0])
        copy_ws = wb.copy_worksheet(ws)
        copy_ws.title = sheet_name
        copy_ws["A2"].value = sheet_name
        copy_ws["A4"].value = values[index][10]
        copy_ws["B7"].value = f"{invoice_month}月分請求書"
        copy_ws["N2"].value = f"{year_month}-{invoice_number:03d}"
        invoice_number += 1
        copy_ws["N3"].value = invoice_date
        copy_ws["A14"].value = values[index][13]
        copy_ws["A15"].value = values[index][18]
        copy_ws["A16"].value = values[index][23]
        copy_ws["A17"].value = values[index][28]
        copy_ws["A18"].value = values[index][33]
        copy_ws["A19"].value = values[index][38]
        copy_ws["A20"].value = values[index][43]
        copy_ws["A21"].value = values[index][48]
        copy_ws["A22"].value = values[index][53]
        copy_ws["A23"].value = values[index][58]
        copy_ws["J14"].value = values[index][14]
        copy_ws["J15"].value = values[index][19]
        copy_ws["J16"].value = values[index][24]
        copy_ws["J17"].value = values[index][29]
        copy_ws["J18"].value = values[index][34]
        copy_ws["J19"].value = values[index][39]
        copy_ws["J20"].value = values[index][44]
        copy_ws["J21"].value = values[index][49]
        copy_ws["J22"].value = values[index][54]
        copy_ws["J23"].value = values[index][59]
        copy_ws["K14"].value = values[index][15]
        copy_ws["K15"].value = values[index][20]
        copy_ws["K16"].value = values[index][25]
        copy_ws["K17"].value = values[index][30]
        copy_ws["K18"].value = values[index][35]
        copy_ws["K19"].value = values[index][40]
        copy_ws["K20"].value = values[index][45]
        copy_ws["K21"].value = values[index][50]
        copy_ws["K22"].value = values[index][55]
        copy_ws["K23"].value = values[index][60]
        copy_ws["L14"].value = values[index][16]
        copy_ws["L15"].value = values[index][21]
        copy_ws["L16"].value = values[index][26]
        copy_ws["L17"].value = values[index][31]
        copy_ws["L18"].value = values[index][36]
        copy_ws["L19"].value = values[index][41]
        copy_ws["L20"].value = values[index][46]
        copy_ws["L21"].value = values[index][51]
        copy_ws["L22"].value = values[index][56]
        copy_ws["L23"].value = values[index][61]
        copy_ws["O14"].value = values[index][17]
        copy_ws["O15"].value = values[index][22]
        copy_ws["O16"].value = values[index][27]
        copy_ws["O17"].value = values[index][32]
        copy_ws["O18"].value = values[index][37]
        copy_ws["O19"].value = values[index][42]
        copy_ws["O20"].value = values[index][47]
        copy_ws["O21"].value = values[index][52]
        copy_ws["O22"].value = values[index][57]
        copy_ws["O23"].value = values[index][62]
        img = Image("files/角印.png")
        img.width = 100
        img.height = 100
        copy_ws.add_image(img, "P5")

ws = wb["請求書"]
wb.remove(ws)

wb.save(output_file)

# ExcelをPDFに変換するための関数
def excel_to_pdf(sheet_name, output_pdf):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False

    wb = excel.Workbooks.Open(os.path.abspath(output_file))
    ws = wb.Sheets(sheet_name)

    # ページ設定を変更して、すべての列が収まるようにする
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesWide = 1
    
    ws.ExportAsFixedFormat(0, os.path.abspath(output_pdf))
    wb.Close(False)
    excel.Application.Quit()

# 各シートごとにPDFを生成
for sheet in wb.sheetnames:
    pdf_file = f"{output_folder}/{sheet}.pdf"
    excel_to_pdf(sheet, pdf_file)

print("各シートごとのPDFが生成されました。")
